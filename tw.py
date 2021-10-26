import time
import openpyxl
import datetime
import calendar

# Получаем книгу
wb = openpyxl.load_workbook("task_support.xlsx")

# Получаем нужный лист
task_sheet = wb['Tasks']

# Получаем полезные переменные с количеством рабочих строк / столбов
rows = task_sheet.max_row
cols = task_sheet.max_column

# Первый таск: Определить количество чётных чисел в столбце 'B'

b_column = task_sheet['B']


def even_numbers(column, rows):
    # Правильный ответ: 507
    i = 0
    for k in range(2, rows):
        cell = column[k].value
        if cell % 2 == 0:
            i += 1
    return f'Всего чётных чисел в столбце "В": {i}'


print(even_numbers(b_column, rows))

# Второй таск: Определить количество простых чисел в столбце 'C'


c_column = task_sheet['C']


def prime_numbers(column, rows):
    # Правильный ответ: 168
    i = 0
    for k in range(2, rows):
        cell = column[k].value
        if cell > 1:
            for j in range(2, column[k].value):
                if (column[k].value % j) == 0:
                    break
            else:
                i += 1
    return f'Всего простых чисел в столбце "С": {i}'


print(prime_numbers(c_column, rows))

# Третий таск: определить количество чисел, которые < 0.5 в столбце 'D'


d_column = task_sheet['D']


def lesser_number(column, rows):
    # Правильный ответ: 485
    i = 0
    for k in range(2, rows):
        cell = column[k].value
        if type(cell) == str:
            cell = float(cell.replace(',', '.').replace(' ', ''))
            if cell < 0.5:
                i += 1
    return f'Всего чисел, которые меньше чем 0.5 в столбце "D": {i}'


print(lesser_number(d_column, rows))

# Четвёртый таск: определить количество вторников в столбце 'E'


e_column = task_sheet['E']


def find_tuesday_by_name(column, rows):
    # Правильный ответ: 156
    i = 0
    for k in range(2, rows):
        cell = column[k].value
        if cell.split()[0] == 'Tue':
            i += 1
    return f'Всего "вторников" в столбце "Е": {i}'


print(find_tuesday_by_name(e_column, rows))

# Пятый таск: определить количество вторников в столбце 'F'

f_column = task_sheet['F']


def find_tuesday_by_date(column, rows):
    # Правильный ответ: 157
    i = 0
    for k in range(2, rows):
        cell = column[k].value
        parsed_date = cell.split()[0].replace('-', '.')
        converted_date = datetime.datetime.strptime(parsed_date, "%Y.%M.%d")
        day_of_the_week = calendar.day_abbr[converted_date.date().weekday()]
        if day_of_the_week == 'Tue':
            i += 1

    return f'Всего вторников в столбце "F": {i}'


print(find_tuesday_by_date(f_column, rows))

# Шестой таск: определить сколько Последних вторников месяца в стобце G:

g_column = task_sheet['G']


def find_final_tuesday_of_the_month(column, rows):
    # Правильный ответ: 43
    i = 0
    for k in range(2, rows):
        cell = column[k].value
        parsed_date = cell.split()[0].replace('-', '.')
        converted_date = datetime.datetime.strptime(parsed_date, "%m.%d.%Y")
        day = converted_date.day
        day_of_the_week = calendar.day_abbr[converted_date.date().weekday()]
        days_in_month = calendar.monthrange(
            converted_date.year, converted_date.month)[1]
        if day_of_the_week == 'Tue' and (day + 7) > days_in_month:
            i += 1

    return f'Всего последних вторников в столбце "G": {i}'


print(find_final_tuesday_of_the_month(g_column, rows))
